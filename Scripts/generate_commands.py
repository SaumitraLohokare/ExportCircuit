# Convert a matrix of files and DER values to exportcircuit command
from openpyxl import Workbook, load_workbook

import sys
import os

if len(sys.argv) != 4:
    print("Format: python generate_commands.py <xlsx file with factors> <folder containing input files> <output folder>")
    print("Folder name containing input files & output folder name must end with '\\'")
    print("To print all commands to a '.bat' file simply add '>commands.bat' as an additional argument.")
    exit()

FEEDER_CELL = 0
CIRCUIT_NAME_CELL = 1
SOLAR_FACTOR_CELL = 5
EV_FACTOR_CELL = 6
HP_FACTOR_CELL = 7

def capital_case(val: str) -> str:
    res = ""
    last_c = " "
    for c in val:
        if last_c == " ":
            res += c.upper()
        else:
            res += c.lower()
        last_c = c

    return res

feeder_to_file = dict()
def fill_feeder_to_file(col: any):
    global feeder_to_file

    feeders = []
    for cell in col:
        feeders.append(str(cell.value))

    for file in os.listdir(in_folder):
        for feeder in feeders:
            if feeder in file:
                feeder_to_file[feeder] = file
    
    print("echo \"Found and filled", len(feeder_to_file), "feeders\"")

def get_command(row: any) -> str:
    feeder = str(row[FEEDER_CELL].value)
    file_name = ""

    if feeder in feeder_to_file.keys():
        file_name = feeder_to_file[feeder]
    else:
        print("echo \"FAILED IN FINDING FILE CORRESPONDING TO FEEDER\"")
        file_name = str(row[CIRCUIT_NAME_CELL].value)
        # Fix for incorrect file names
        # if file_name in incorrect_file_names.keys():
        #     file_name = incorrect_file_names[file_name]
        file_name = capital_case(file_name.removeprefix("NYS-")) + ".xlsx"

    folder_file = in_folder + file_name
    folder_file_has_space = False
    for c in folder_file:
        folder_file_has_space |= c == " "

    if folder_file_has_space:
        in_value = "-in \"" + folder_file+ "\""
    else:
        in_value = "-in " + folder_file

    out_value = ""
    if out_has_spaces:
        out_value = "-out \"" + out_folder + "\""
    else:
        out_value = "-out " + out_folder

    solar_factor = str(row[SOLAR_FACTOR_CELL].value)
    solar_value = "-addsolar " + solar_factor

    ev_factor = str(row[EV_FACTOR_CELL].value)
    ev_value = "-addevchargingstation " + ev_factor

    hp_factor = str(row[HP_FACTOR_CELL].value)
    hp_value = "-addheatpump " + hp_factor

    return "exportcircuit" + " " + in_value + " " + out_value + " " + solar_value + " " + ev_value + " " + hp_value

if __name__ == "__main__":
    in_file = sys.argv[1]
    in_folder = sys.argv[2]
    out_folder = sys.argv[3]
    out_has_spaces = False
    for c in out_folder:
        out_has_spaces |= c == " "

    book = load_workbook(in_file)
    sheet = book.active

    last_row = sheet.max_row
    rows = sheet.rows

    feeder_col = list(sheet.columns)[FEEDER_CELL]
    fill_feeder_to_file(feeder_col)

    header_row = True
    for row in rows:
        # Skip the first row
        if header_row: 
            header_row = False
            continue

        print(get_command(row))

# incorrect_file_names = {
#     "NYS-ENDICOTT CLARK ST 734": "NYS-ENDICOTT CLARK 734",
#     "NYS-ENDICOTT CLARK ST 735": "NYS-ENDICOTT CLARK 735",
#     "NYS-GENEGANTSLET CORNERS 422": "NYS-GENEGANTSLET COR 422",
#     "NYS-GLEN AUBREY 417": "NYS-GLEN AUBRY 417",
#     "NYS-WHITNEY POINT 780": "NYS-WHITNEY PT 780",
#     "NYS-WHITNEY POINT 781": "NYS-WHITNEY PT 781",
#     "NYS-WHITNEY POINT 782": "NYS-WHITNEY PT 782",
#     "NYS-VINCENT CORNERS 269": "NYS-VINCENT CORNERS",
#     "NYS-RANO BLVD 694": "NYS-RANO 694",
#     "NYS-NOWLAN RD 226": "NYS-NOWLAN 226",
#     "NYS-NOWLAN RD 227": "NYS-NOWLAN 227",
#     "NYS-NOWLAN RD 228": "NYS-NOWLAN 228",
#     "NYS-MORRIS ST 657": "NYS-MORRIS 657",
#     "NYS-MORRIS ST 658": "NYS-MORRIS 658",
#     "NYS-MORRIS ST 659": "NYS-MORRIS 659",
#     "NYS-HOOPER RD 701": "NYS-HOOPER 701",
#     "NYS-HOOPER RD 702": "NYS-HOOPER 702",
#     "NYS-HOOPER RD 703": "NYS-HOOPER 703",
#     "NYS-HOOPER RD 704": "NYS-HOOPER 704",
#     "NYS-JARVIS ST 687": "NYS-JARVIS 687",
#     "NYS-JARVIS ST 689": "NYS-JARVIS 689",
#     "NYS-JARVIS ST 690": "NYS-JARVIS 690",
#     "NYS-JARVIS ST 691": "NYS-JARVIS 691",
#     "NYS-JARVIS ST 692": "NYS-JARVIS 692",
#     "NYS-KATTLEVILLE 422": "NYS-KATTELVILLE 422"
# }

# def get_command(row: any) -> str:
#     file_name = str(row[1].value)
#     # Fix for incorrect file names
#     if file_name in incorrect_file_names.keys():
#         file_name = incorrect_file_names[file_name]
#     file_name = capital_case(file_name.removeprefix("NYS-"))

#     in_value = "-in \"" + in_folder + file_name + " " + str(row[0].value) + ".xlsx\""
#     out_value = ""
#     if out_has_spaces:
#         out_value = "-out \"" + out_folder + "\""
#     else:
#         out_value = "-out " + out_folder

#     solar_factor = str(row[SOLAR_FACTOR_CELL].value)
#     solar_value = "-addsolar " + solar_factor

#     ev_factor = str(row[EV_FACTOR_CELL].value)
#     ev_value = "-addevchargingstation " + ev_factor

#     hp_factor = str(row[HP_FACTOR_CELL].value)
#     hp_value = "-addheatpump " + hp_factor

#     return "exportcircuit" + " " + in_value + " " + out_value + " " + solar_value + " " + ev_value + " " + hp_value